"""Minimal script to convert Excel XLSX format to CSV.
"""
import argparse
import csv
from openpyxl import load_workbook

if __name__ == '__main__':
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("--xlsx", required=True, help="path to XLSX file to convert")
    parser.add_argument("--output", required=True, help="path to CSV output")

    args = parser.parse_args()

    workbook = load_workbook(filename=args.xlsx)
    sheet = workbook.active
    field_names = [cell.value for cell in sheet[1]]

    with open(args.output, "w", encoding="utf-8", newline='') as csvfile:
        writer = csv.writer(csvfile, dialect="unix")
        writer.writerow(field_names)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            writer.writerow(row)
